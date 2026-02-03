import os
from datetime import datetime #, date
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries
import shutil


# =========================
# CONFIG (paths + filenames)
# =========================

onedrive_root = os.path.join(
    os.path.expanduser("~"),
    "OneDrive - Luis Domingo Trucco S.A")

INPUT_DIR = os.path.join(onedrive_root, "outputs", "Informes Roic", "Renta Fija")

# INPUT_DIR = r"C:\Users\desar\OneDrive - Luis Domingo Trucco S.A\outputs\Informes Roic\Renta Fija"
# INPUT_DIR = r"C:\Users\desar\OneDrive\Desktop\LUCAS ROIC\Informes bonos"

OUTPUT_DIR = os.path.join(onedrive_root, "inputs", "Cashflows")

# OUTPUT_DIR = r"C:\Users\desar\OneDrive - Luis Domingo Trucco S.A\inputs\Cashflows"

OUTPUT_FILE = os.path.join(OUTPUT_DIR, "Cashflows.xlsx")

# Ajustar los nombres si los archivos difieren
WB_SOBERANOS = os.path.join(INPUT_DIR, "BONOS SOBERANOS Y BOPREALES EN DOLARES.xlsm")
WB_ONS = os.path.join(INPUT_DIR, "ONs EN DOLARES.xlsm")
WB_LETRAS = os.path.join(INPUT_DIR, "LETRAS EN PESOS.xlsm")
WB_PESOS = os.path.join(INPUT_DIR, "BONOS EN PESOS.xlsm")

PLANTILLA_DIR = os.path.join(OUTPUT_DIR, "viejos y plantilla")
PLANTILLA_FILE = os.path.join(PLANTILLA_DIR, "Plantilla.xlsx")

VIEJOS_DIR = PLANTILLA_DIR
# =========================
# Helpers básicos
# =========================
def _cell_value(cell):
    return cell.value

def _as_date(x):
    # openpyxl suele devolver datetime/date ya listos
    return x


def read_tir_sheet_ab(wb_path, sheet_name="TIR"):
    wb = load_workbook(wb_path, data_only=True, read_only=True)
    ws = wb[sheet_name]

    rows = []
    # Columnas A y B: desde fila 1 hacia abajo hasta que A esté vacío
    r = 1
    while True:
        a = _cell_value(ws.cell(row=r, column=1))
        b = _cell_value(ws.cell(row=r, column=2))
        if a is None or a == "":
            break
        rows.append((a, b))
        r += 1

    wb.close()
    return pd.DataFrame(rows, columns=["Ticker", "TIR"])


def read_table_until_blank(ws, header_row=1, first_data_row=2, stop_col=1):
    """
    Lee una tabla simple desde header_row. Termina cuando la columna stop_col (por defecto B) esté vacía.
    No hace validaciones extra.
    """
    headers = []
    c = 1
    while True:
        h = _cell_value(ws.cell(row=header_row, column=c))
        if h is None or h == "":
            break
        headers.append(h)
        c += 1

    data = []
    r = first_data_row
    while True:
        stop_val = _cell_value(ws.cell(row=r, column=stop_col))
        if stop_val is None or stop_val == "":
            break
        row = []
        for j in range(1, len(headers) + 1):
            row.append(_cell_value(ws.cell(row=r, column=j)))
        data.append(row)
        r += 1

    return pd.DataFrame(data, columns=headers)


def _find_header_col_idx(headers, wanted_list):
    """
    Devuelve el índice (0-based) del primer header que coincida exactamente con alguno de wanted_list.
    """
    for w in wanted_list:
        if w in headers:
            return headers.index(w)
    return None


def moneda_from_clasificacion(clasif):
    
    c = str(clasif)
    
    out = 'ARS'
    
    if c.startswith(("Bop", "Sob", "ON")) and "DDL" not in c:
        out = "USD"
        
    return out

def parse_cashflow_common_or_special(ws, ticker):
    """
    Caso 2 o 3 (versión robusta a A3 vacío):
    - Headers: fila 3, desde columna D (col=4) hacia la derecha.
    - Datos: desde fila 5 hacia abajo.
    - Corta cuando VENCIMIENTO esté vacío.
    - Si existe 'INTERESES CF' en headers -> usa eso para Intereses.
      Si no, recrea Intereses. como amortización y Flujo = Intereses + Amortización.
    """

    header_row = 3
    data_row = 5
    start_col = 4  # D

    def _to_num(x):
        # Convierte valores típicos de Excel a float sin ponerse creativo
        if x is None or x == "" or x == "-":
            return 0.0
        return float(x)

    # 1) Leer toda la fila de headers desde D3 hasta max_column
    #    (usando iter_rows para NO ir celda por celda)
    max_col = ws.max_column
    header_vals = next(ws.iter_rows(
        min_row=header_row, max_row=header_row,
        min_col=start_col, max_col=max_col,
        values_only=True
    ))

    # Si hay trailing Nones, no importa; armamos lista completa tal cual
    headers = list(header_vals)

    # 2) Mapear nombres de columnas a índices (0-based dentro de este "bloque" desde col D)
    #    Ojo: el user pidió nombres exactos en MAYÚSCULAS.
    def _idx(name):
        try:
            return headers.index(name)
        except ValueError:
            return None

    idx_venc = _idx("VENCIMIENTO")
    idx_cupon = _idx("CUPON %")
    idx_resid = _idx("RESIDUAL")
    idx_amort = _idx("AMORTIZ.")
    # Estos dos pueden o no estar:
    idx_int_cf = _idx("INTERESES CF")

    # En tu estructura, estas 4 deberían existir siempre (según tu regla)
    # Si alguna no está, va a explotar más adelante (como preferís: keep it simple).

    # 3) Leer datos desde fila 5 hacia abajo, desde col D hacia max_col
    rows = []
    prev_date = None

    for row_vals in ws.iter_rows(
        min_row=data_row, max_row=ws.max_row,
        min_col=start_col, max_col=max_col,
        values_only=True
    ):
        venc = row_vals[idx_venc]  # fecha
        if venc is None or venc == "":
            break

        cupon = row_vals[idx_cupon]
        resid = row_vals[idx_resid]
        amort = row_vals[idx_amort]

        # Normalizar numéricos básicos (sin adivinar formatos)
        cupon_n = _to_num(cupon)
        resid_n = _to_num(resid)
        amort_n = _to_num(amort)
        
        
        # Si existe INTERESES CF, lo usamos; si no, calculamos
        if idx_int_cf is not None:
            inteerses_n = _to_num(row_vals[idx_int_cf])
        else:
            # recrear intreses con ACT/360 simple
            if prev_date is None:
                intereses_n = 0.0
            else:
                dt_days = (venc- prev_date).days
                intereses_n = (resid_n * cupon_n * dt_days) / 360
                
        # flujo SIEMPRE = intereses + amortización
        flujo_n = intereses_n + amort_n

        rows.append([ticker, venc, cupon_n, resid_n, intereses_n, amort_n, flujo_n])
        prev_date = venc

    df = pd.DataFrame(
        rows,
        columns=["Ticker", "Fecha", "Cupón", "Residual", "Intereses", "Amortización", "Flujo"]
    )
    return df


def build_letras_cashflows_from_lecaps(ws):
    """
    LETRAS EN PESOS:
    hoja LECAPS con headers fila 1.
    Se corta cuando Ticker esté vacío (fila en blanco).
    Genera 2 filas por ticker:
      - Fecha 1 = Emisión: Flujo=0; cupón/residual/intereses/amort en blanco
      - Fecha 2 = Vencimiento: Residual=100; Intereses=0; Amortización=100; Flujo=Monto al Vto
    """
    # Leemos headers explícitos tal como vienen (fila 1)
    df = read_table_until_blank(ws, header_row=1, first_data_row=2, stop_col=1)  # stop en columna A (Ticker)

    # Columnas relevantes (nombres exactos de tu hoja)
    col_ticker = "Ticker"
    col_emision = "Emisión"
    col_venc = "Vencimiento"
    col_monto_vto = "Monto al Vto"
    col_tir = "TIR"

    # TIR table (para reemplazar hoja TIR)
    df_tir = df[[col_ticker, col_tir]].copy()
    df_tir.columns = ["Ticker", "TIR"]

    rows_cf = []
    for _, row in df.iterrows():
        t = row[col_ticker]
        f0 = row[col_emision]
        f1 = row[col_venc]
        monto = row[col_monto_vto]

        # Fila emisión
        rows_cf.append([t, f0, "", "", "", "", 0])

        # Fila vencimiento
        rows_cf.append([t, f1, "", 100, 0, 100, monto])

    df_cf = pd.DataFrame(
        rows_cf,
        columns=["Ticker", "Fecha", "Cupón", "Residual", "Intereses", "Amortización", "Flujo"]
    )

    return df_tir, df_cf


def build_bonos_tabla_tipo_letras(ws, tickers_filter):
    """
    Para BONOS EN PESOS: hojas Bonos_DDL / Bonos_CER / Bonos_TAMAR
    Se leen "igual que LETRAS EN PESOS" (tabla plana) y luego se filtra por tickers de hoja TIR.
    Genera 2 filas por ticker (Emisión y Vencimiento) igual que letras.
    """
    df = read_table_until_blank(ws, header_row=1, first_data_row=2, stop_col=1)

    col_ticker = "Ticker"
    col_emision = "Emisión"
    col_venc = "Vencimiento"
    col_monto_vto = "Monto al Vto"

    # Filtrar por tickers de la hoja TIR del libro BONOS EN PESOS
    df = df[df[col_ticker].isin(tickers_filter)].copy()

    rows_cf = []
    for _, row in df.iterrows():
        t = row[col_ticker]
        f0 = row[col_emision]
        f1 = row[col_venc]
        monto = row[col_monto_vto]

        rows_cf.append([t, f0, "", "", "", "", 0])
        rows_cf.append([t, f1, "", 100, 0, 100, monto])

    df_cf = pd.DataFrame(
        rows_cf,
        columns=["Ticker", "Fecha", "Cupón", "Residual", "Intereses", "Amortización", "Flujo"]
    )
    return df_cf


# =========================
# Clasificación
# =========================
def clasif_soberanos_bopreales(ticker):
    if isinstance(ticker, str) and len(ticker) > 0:
        if ticker[0] == "B":
            return "Bopreal"
        if ticker[0] == "G":
            return "Sob. Ley Extranjera"
    return "Sob. Ley Local"


def clasif_bonos_pesos(ticker, source_bucket, source_sheet=None):
    """
    source_bucket:
      - "INDIV" (caso 2/3 hoja individual)
      - "Bonos_CER"
      - "Bonos_DDL"
      - "Bonos_TAMAR"
    """
    if source_bucket == "INDIV":
        if ticker == "TO26":
            return "Bono PV"
        return "CER"
    if source_bucket == "Bonos_CER":
        return "CER"
    if source_bucket == "Bonos_DDL":
        return "Dollar linked"
    if source_bucket == "Bonos_TAMAR":
        if isinstance(ticker, str) and ticker.startswith("M"):
            return "TAMAR +Margen"
        return "TAMAR con Tasa Fija"
    return ""

def clasif_ons_por_listas(ticker, set_leg_arg, set_leg_eeuu, set_dl):
    """
    Clasifica por pertenencia a listas (por construcción: una y solo una).
    """
    if ticker in set_leg_arg:
        return "ON Ley Local"
    if ticker in set_leg_eeuu:
        return "ON Ley Extr."
    if ticker in set_dl:
        return "ON DDL"
    return ""

def clasif_letras_por_prefijo(ticker):
    if ticker is None or ticker == "":
        return "LETRA"
    t = str(ticker)
    if t.startswith("S"):
        return "LECAP"
    if t.startswith("T"):
        return "BONCAP"
    return "LETRA"

def _read_defined_name_set(wb, def_name):
    """
    Lee un nombre definido (Named Range) y devuelve un set con los valores no vacíos.
    Asume que el rango contiene una columna de tickers (puede ser 1 columna o más, tomamos todo).
    """
    dn = wb.defined_names.get(def_name)
    if dn is None:
        return set()

    values = []
    # Un nombre definido puede apuntar a uno o varios destinos (sheet, range)
    for title, coord in dn.destinations:
        ws = wb[title]

        # coord puede ser tipo "$A$1:$A$50" o "A1:A50"
        min_col, min_row, max_col, max_row = range_boundaries(coord)

        for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                                min_col=min_col, max_col=max_col,
                                values_only=True):
            for v in row:
                if v is None or v == "":
                    continue
                values.append(v)

    # Importantísimo: no normalizar “por las dudas”; solo strip si es string para evitar espacios
    out = set()
    for v in values:
        if isinstance(v, str):
            v = v.strip()
        out.add(v)
    return out





def clear_range_values(ws, start_row, start_col, end_row, end_col):
    """
    Limpia valores (pone None) en un rango, sin tocar formato ni tabla.
    """
    for row in ws.iter_rows(min_row=start_row, max_row=end_row,
                            min_col=start_col, max_col=end_col):
        for cell in row:
            cell.value = None


def write_df_into_fixed_block(ws, df, start_row=2, start_col=1, max_cols=8):
    """
    Escribe df en un bloque fijo (por ej A:H) empezando en start_row.
    No toca headers ni tablas. Solo valores.
    """
    nrows = df.shape[0]
    ncols = df.shape[1]

    # Keep it simple: asumimos df tiene <= max_cols columnas y están en el orden correcto
    for i, row in enumerate(df.itertuples(index=False), start=start_row):
        for j, val in enumerate(row, start=start_col):
            ws.cell(row=i, column=j, value=val)



def archive_existing_output(output_file, viejos_dir):
    """
    Si output_file existe:
      - mueve a viejos_dir
      - nombre: cashflows_DD-MM-YYYY.xlsx
      - si existe, cashflows_DD-MM-YYYY.2.xlsx, .3, ...
    """
    if not os.path.exists(output_file):
        return None

    os.makedirs(viejos_dir, exist_ok=True)

    today_str = datetime.today().strftime("%d-%m-%Y")
    base_name = f"cashflows_{today_str}"
    ext = ".xlsx"

    candidate = os.path.join(viejos_dir, base_name + ext)
    if not os.path.exists(candidate):
        target = candidate
    else:
        k = 2
        while True:
            candidate_k = os.path.join(viejos_dir, f"{base_name}.{k}{ext}")
            if not os.path.exists(candidate_k):
                target = candidate_k
                break
            k += 1

    shutil.move(output_file, target)
    return target

def create_output_from_template(plantilla_file, output_file):
    """
    Copia Plantilla.xlsx a Cashflows.xlsx.
    """
    shutil.copy2(plantilla_file, output_file)

#%% =========================
# MAIN
# =========================
def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    # --------
    # 1) TIRs
    # --------
    df_tir_sober = read_tir_sheet_ab(WB_SOBERANOS, "TIR")
    df_tir_ons = read_tir_sheet_ab(WB_ONS, "TIR")
    df_tir_pesos = read_tir_sheet_ab(WB_PESOS, "TIR")

    # Letras (LECAPS)
    wb_letras = load_workbook(WB_LETRAS, data_only=True, read_only=True)
    ws_lecaps = wb_letras["LECAPS"]
    df_tir_letras, df_cf_letras = build_letras_cashflows_from_lecaps(ws_lecaps)
    wb_letras.close()

    # Concatenar TIR total (4 fuentes)
    df_tir_total = pd.concat([df_tir_sober, df_tir_ons, df_tir_pesos, df_tir_letras], ignore_index=True)

    # Sets de tickers por libro (para filtrar cashflows)
    tickers_sober = set(df_tir_sober["Ticker"].dropna().tolist())
    tickers_ons = set(df_tir_ons["Ticker"].dropna().tolist())
    tickers_pesos = set(df_tir_pesos["Ticker"].dropna().tolist())

    # -----------------------
    # 2) Cashflows por libro
    # -----------------------
    cashflows_all = []

    # 2A) LETRAS EN PESOS (ya construido)
    df_cf_letras = df_cf_letras.copy()
    df_cf_letras["Clasificación"] = df_cf_letras["Ticker"].apply(clasif_letras_por_prefijo)
    df_cf_letras["Moneda"] = df_cf_letras["Clasificación"].apply(moneda_from_clasificacion)
    cashflows_all.append(df_cf_letras[["Ticker", "Clasificación", "Fecha", "Cupón", "Residual", "Intereses", "Amortización", "Flujo"]])

    # # 2B) BONOS SOBERANOS Y BOPREALES EN DOLARES (caso 2/3 por hoja)
    wb_sober = load_workbook(WB_SOBERANOS, data_only=True, read_only=True)
    for t in tickers_sober:
        ws = wb_sober[t]
        df_cf = parse_cashflow_common_or_special(ws, t)
        df_cf["Clasificación"] = clasif_soberanos_bopreales(t)
        
        cashflows_all.append(df_cf[["Ticker", "Clasificación", "Fecha", "Cupón", "Residual", "Intereses", "Amortización", "Flujo"]])
    wb_sober.close()

    # 2C) ONs EN DOLARES (caso 2/3 por hoja + clasificación por hex tab)
    wb_ons = load_workbook(WB_ONS, data_only=True, read_only=True)
    
    # ONs: workbook normal (NO read_only) para leer defined names
    wb_ons_meta = load_workbook(WB_ONS, data_only=True, read_only=False)
    
    set_leg_arg  = _read_defined_name_set(wb_ons_meta, "leg_arg")
    set_leg_eeuu = _read_defined_name_set(wb_ons_meta, "leg_eeuu")
    set_dl       = _read_defined_name_set(wb_ons_meta, "dl")

    for t in tickers_ons:
        ws = wb_ons[t]
        # ws_meta = wb_ons_meta[t]
        # print("DEBUG t (repr):", repr(t), "type:", type(t))
        # print("DEBUG sheets count:", len(wb_ons_meta.sheetnames))
        # print("DEBUG first 10 sheets:", wb_ons_meta.sheetnames[:10])
        # print("DEBUG t in sheetnames?:", t in wb_ons_meta.sheetnames)
        df_cf = parse_cashflow_common_or_special(ws, t)
        
        df_cf["Clasificación"] = clasif_ons_por_listas(t, set_leg_arg, set_leg_eeuu, set_dl)
        
        
        cashflows_all.append(df_cf[["Ticker", "Clasificación", "Fecha", "Cupón", "Residual", "Intereses", "Amortización", "Flujo"]])
    
    wb_ons.close()
    wb_ons_meta.close()
    # 2D) BONOS EN PESOS
    
    # print("BONOS EN PESOS SECTOR")
    wb_pesos = load_workbook(WB_PESOS, data_only=True, read_only=True)

    # 2D-1) Hojas tipo letras: Bonos_DDL / Bonos_CER / Bonos_TAMAR
    for sh in ["Bonos_DDL", "Bonos_CER", "Bonos_TAMAR"]:
        ws = wb_pesos[sh]
        df_cf = build_bonos_tabla_tipo_letras(ws, tickers_pesos)
        df_cf["Clasificación"] = df_cf["Ticker"].apply(lambda x, s=sh: clasif_bonos_pesos(x, s))
        
        cashflows_all.append(df_cf[["Ticker", "Clasificación", "Fecha", "Cupón", "Residual", "Intereses", "Amortización", "Flujo"]])

    # 2D-2) Tickes restantes (no están en esas 3 hojas) => hoja individual caso 2/3
    # Primero armamos set de tickers ya cubiertos por Bonos_*
    covered = set()
    for sh in ["Bonos_DDL", "Bonos_CER", "Bonos_TAMAR"]:
        ws = wb_pesos[sh]
        df_tmp = read_table_until_blank(ws, header_row=1, first_data_row=2, stop_col=2)
        covered |= set(df_tmp["Ticker"].dropna().tolist())
    remaining = [t for t in tickers_pesos if t not in covered]

    for t in remaining:
        ws = wb_pesos[t]
        df_cf = parse_cashflow_common_or_special(ws, t)
        df_cf["Clasificación"] = clasif_bonos_pesos(t, "INDIV")
        
        cashflows_all.append(df_cf[["Ticker", "Clasificación", "Fecha", "Cupón", "Residual", "Intereses", "Amortización", "Flujo"]])

    wb_pesos.close()

    # Consolidado final
    df_cashflows = pd.concat(cashflows_all, ignore_index=True)    
    df_cashflows["Moneda"] = df_cashflows["Clasificación"].apply(moneda_from_clasificacion)
    
    # ==========================
    # OUTPUT: plantilla + viejos
    # ==========================
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    os.makedirs(PLANTILLA_DIR, exist_ok=True)
    os.makedirs(VIEJOS_DIR, exist_ok=True)

    # 1) Archivar Cashflows.xlsx actual si existe
    archive_existing_output(OUTPUT_FILE, VIEJOS_DIR)

    # 2) Crear el nuevo Cashflows.xlsx copiando plantilla
    create_output_from_template(PLANTILLA_FILE, OUTPUT_FILE)

    # 3) Abrir el nuevo output y escribir data (manteniendo formato)
    wb_out = load_workbook(OUTPUT_FILE)

    ws_cf = wb_out["CASHFLOWS"]
    ws_tir = wb_out["TIR"]

    # Limpieza + escritura (tu enfoque de bloque fijo A:H y A:B)
    clear_range_values(ws_cf, start_row=2, start_col=1, end_row=50000, end_col=9)
    clear_range_values(ws_tir, start_row=2, start_col=1, end_row=50000, end_col=2)

    write_df_into_fixed_block(ws_cf, df_cashflows, start_row=2, start_col=1, max_cols=9)
    write_df_into_fixed_block(ws_tir, df_tir_total, start_row=2, start_col=1, max_cols=2)

    wb_out.save(OUTPUT_FILE)
    wb_out.close()

if __name__ == "__main__":
    main()


